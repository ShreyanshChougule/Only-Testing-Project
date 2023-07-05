import openpyxl


class Excel_File_Opractions:
    Path = "C:\\Users\\Tejas\\Only Testing Project\\TestData\\Test Data.xlsx"

    def read_excel(self, rownum, colnum):
        Excel = openpyxl.load_workbook(self.Path)
        Sheet = Excel.active
        return Sheet.cell(row=rownum, column=colnum).value

    def write_excel(self, rownum, colnum, data):
        Book = openpyxl.load_workbook(self.Path)
        Sheet = Book.active
        Sheet.cell(row=rownum, column=colnum).value = data
        Book.save(self.Path)
